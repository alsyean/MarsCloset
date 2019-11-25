from openpyxl import load_workbook
from operator import eq


#엑셀 불러오기
load_cloth=load_workbook("C:\Users\pyj\Desktop\My Node Js 1005/python/vm.xlsx", data_only=True)
cloth_excel=load_cloth['cloth']
style_excel=load_cloth['style']

#pre = 레코드, pnum = 레코드 수
def style(pre, row):
    # 선택한 선호 스타일의 수만큼 for문을 돌리고 해당하는 옷을 가져옴

    if pre[row][2] == "Lovely":
        #Top
        sqlT = " category LIKE 'top' "
        #Bottom
        sqlB = " category LIKE 'bottoms' and subclass1 LIKE 'skirt'"

    elif pre[row][2] == "Dandy":
        sqlT = " category LIKE 'top' and subclass3 NOT IN ('crop', 'hoody', 'halter', 'casual')"
        sqlB = " category LIKE 'bottoms' and subclass1 IN ('skirt', 'pants')"
    
    elif pre[row][2] == "casual":
        sqlT = " category LIKE 'top' and subclass3 NOT IN ('mantoman', 'hody', 'business', 'casual')"
        #sqlB = " category LIKE 'bottom' and subclass1 LIKE 'skirt'"
        
    else:
        sqlT = " category LIKE 'top' and subclass3 NOT IN ('mantoman', 'hody', 'business', 'casual')"
        #sqlB = " category LIKE 'bottom' and subclass1 LIKE 'skirt'"

    return sqlT, sqlB        
    

#계절
#clothes_season : 옷의 시즌을 구별할 쿼리문을 담을 변수

def season(temp) :
    if temp <= 2:
        #겨울
        clothes_season = " and season LIKE 'W'"
    elif temp > 2 and temp <= 4:
        #겨울 봄
        clothes_season = " and season regexp 'W|S'"
    elif temp > 5 and temp <= 19:
        #봄 가을
        clothes_season= " and season regexp 'S|F"
    elif temp > 19 and temp <= 21:
        #봄 가을 여름
        clothes_season = " and season regexp 'S|F|U$'" 
    elif temp > 21 and temp <=24:
        #가을 여름
        clothes_season = " and season regexp 'F|U$'"
    else:
        #여름
        clothes_season =  " and season LIKE 'U'"

    return clothes_season
    

#사용자정의함수 vm사용
def usedvm(bottoms1,bottoms2,bottoms3,bottoms4,bottomsstyle,top1,top2,top3,top4,topstyle) :
   
    #//////////////////////////////하의
    ###############################pants
    if eq(bottoms1,"pants"):
        bottoms=5
    ###############################
        if eq(bottoms2,"long"):
            bottoms+=0
        elif eq(bottoms2,"short"):
            bottoms+=40
    ###############################
        if eq(bottoms3,"bootcut"):
            bottoms+=0
        elif eq(bottoms3, "cagopants"):
            bottoms+=8
        elif eq(bottoms3, "jeans"):
            bottoms+=16
        elif eq(bottoms3, "jogger"):
            bottoms+=24
        elif eq(bottoms3, "leggings"):
            bottoms+=32
    #################################
        if eq(bottoms4,"printing"):
            bottoms+=0
        elif eq(bottoms4, "dot"):
            bottoms+=1
        elif eq(bottoms4, "leopard"):
            bottoms+=2
        elif eq(bottoms4, "race"):
            bottoms+=3
        elif eq(bottoms4, "basic"):
            bottoms+=4
        elif eq(bottoms4, "stripe"):
            bottoms+=5
        elif eq(bottoms4, "check"):
            bottoms+=6
        elif eq(bottoms4, "camouflage"):
            bottoms+=7
    ###################################@@@@@@@@@@@@@@@@
            
    ###################################skirt
    elif eq(bottoms1,"skirt") and eq(bottoms2,"skirt"):
        bottoms=85
    ###############################
        if eq(bottoms3,"accordion skirt"):
            bottoms+=0
        elif eq(bottoms3, "a-line skirt"):
            bottoms+=8
        elif eq(bottoms3, "mini skirt"):
            bottoms+=16
        elif eq(bottoms3, "tube skirt"):
            bottoms+=24
        elif eq(bottoms3, "wrap skirt"):
            bottoms+=32
    #################################
        if eq(bottoms4,"printing"):
            bottoms+=0
        elif eq(bottoms4, "dot"):
            bottoms+=1
        elif eq(bottoms4, "leopard"):
            bottoms+=2
        elif eq(bottoms4, "race"):
            bottoms+=3
        elif eq(bottoms4, "basic"):
            bottoms+=4
        elif eq(bottoms4, "stripe"):
            bottoms+=5
        elif eq(bottoms4, "check"):
            bottoms+=6
        elif eq(bottoms4, "camouflage"):
            bottoms+=7

    ###################################@@@@@@@@@@@@@@@@
    #//////////////////////////////하의


            
    #//////////////////////////////상의
    ##################################t-shirts
    if eq(top1,"t-shirts"):
        top=6
        if eq(top2,"logn sleeve"):
            top+=0
        elif eq(top2,"short sleeve"):
            top+=48
        elif eq(top2,"sleeveless"):
            top+=96
    #################################
        if eq(top3,"box"):
            top+=0
        elif eq(top3,"crop"):
            top+=8
        elif eq(top3,"halter"):
            top+=16
        elif eq(top3,"hoody"):
            top+=24
        elif eq(top3,"mantoman"):
            top+=32
        elif eq(top3,"neat"):
            top+=40

    #################################
        if eq(top4,"printing"):
            top+=0
        elif eq(top4,"dot"):
            top+=1
        elif eq(top4,"leopard"):
            top+=2
        elif eq(top4,"race"):
            top+=3
        elif eq(top4,"basic"):
            top+=4
        elif eq(top4,"stripe"):
            top+=5
        elif eq(top4,"check"):
            top+=6
        elif eq(top4,"camouflage"):
            top+=7
    ##################################@@@@@@@@@@@@@@@@
            
    ##################################shirts
    elif eq(top1,"shirts"):
        top=150
        if eq(top2,"long"):
            top+=0
        elif eq(top2,"short"):
            top+=16
    #################################
        if eq(top3,"business"):
            top+=0
        elif eq(top3,"casual"):
            top+=8
    #################################
        if eq(top4,"printing"):
            top+=0
        elif eq(top4,"dot"):
            top+=1
        elif eq(top4,"leopard"):
            top+=2
        elif eq(top4,"race"):
            top+=3
        elif eq(top4,"basic"):
            top+=4
        elif eq(top4,"stripe"):
            top+=5
        elif eq(top4,"check"):
            top+=6
        elif eq(top4,"camouflage"):
            top+=7
    ##################################@@@@@@@@@@@@@@@@
    ##################################outer
    elif eq(top1,"outer") and eq(top2,"outer"):
        top=182
    #################################
        if eq(top3,"cardigun"):
            top+=0
        elif eq(top3,"coat"):
            top+=8
        elif eq(top3,"hoodyzipup"):
            top+=16
        elif eq(top3,"jacket"):
            top+=24
        elif eq(top3,"padding"):
            top+=32
        elif eq(top3,"windbreaker"):
            top+=40
    #################################
        if eq(top4,"printing"):
            top+=0
        elif eq(top4,"dot"):
            top+=1
        elif eq(top4,"leopard"):
            top+=2
        elif eq(top4,"race"):
            top+=3
        elif eq(top4,"basic"):
            top+=4
        elif eq(top4,"stripe"):
            top+=5
        elif eq(top4,"check"):
            top+=6
        elif eq(top4,"camouflage"):
            top+=7
    ##################################@@@@@@@@@@@@@@@@

    ##################################onepiece
    elif eq(top1, "onepiece"):
        top=230
        bottoms=5
    ##################################@@@@@@@@@@@@@@@
    #//////////////////////////////상의





    ######@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@style
    global tstyle, bstyle
    #top
    if eq(topstyle,"colorful"):
        tstyle=3
    elif eq(topstyle,"lucid"):
        tstyle=4
    elif eq(topstyle,"natural"):
        tstyle=5
    elif eq(topstyle,"stiff"):
        tstyle=6
    elif eq(topstyle,"strong"):
        tstyle=7
    elif eq(topstyle,"warm"):
        tstyle=8
    elif eq(topstyle,"zingzy"):
        tstyle=9

    #bottoms

    if eq(bottomsstyle,"colorful"):
        bstyle=3
    elif eq(bottomsstyle,"lucid"):
        bstyle=4
    elif eq(bottomsstyle,"natural"):
        bstyle=5
    elif eq(bottomsstyle,"stiff"):
        bstyle=6
    elif eq(bottomsstyle,"strong"):
        bstyle=7
    elif eq(bottomsstyle,"warm"):
        bstyle=8
    elif eq(bottomsstyle,"zingzy"):
        bstyle=9

    if eq(top1, "onepiece"):
        stylepoint=1
    else :
        stylepoint=style_excel.cell(tstyle,bstyle).value

    clothpoint=cloth_excel.cell(top,bottoms).value
    allpoint=stylepoint+clothpoint
    
    return allpoint

#//////////////////////////////////////////////////////////



